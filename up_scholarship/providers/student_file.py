from openpyxl import load_workbook
from collections import OrderedDict
import json
from up_scholarship.providers.constants import StudentFileTypes


class StudentFile:
	def read_file(self, filename, file_type):
		"""read the appropriate file using the file type"""
		if file_type == StudentFileTypes.json:
			return self._readjson(filename)
		elif file_type == StudentFileTypes.excel:
			return self._readexcel(filename)
		else:
			return None

	def write_file(self, students, infile, outfile, type):
		"""write the appropriate file using the file type"""
		if type == StudentFileTypes.json:
			return self._writejson(students, outfile)
		elif type == StudentFileTypes.excel:
			return self._writeexcel(students, infile, outfile)
		else:
			return None

	def _readexcel(self, filename: str) -> list:
		"""read excel file assuming headers as keys and return dict list"""
		wb = load_workbook(filename=filename, read_only=True, data_only=True)
		ws = wb.active
		i_row = 0
		i_column = 0
		headers_list = []
		students = []
		for row in ws.rows:
			student = OrderedDict()
			for cell in row:
				if i_row == 0:
					headers_list.append(cell.value)
				else:
					value = cell.value
					if value is None:
						value = ''
					student[headers_list[i_column]] = str(value)
					i_column += 1
			if i_row != 0:
				students.append(student)
			i_column = 0
			i_row += 1
		return students

	def _readjson(self, filename: str) -> list:
		"""read file as json and return dict list"""
		with open(filename, 'r') as fp:
			js = json.load(fp)
		return js

	def _writeexcel(self, students: list, in_filename: str, out_filename: str):
		"""write dict list to excel by assuming first element as headers"""
		headers_list = students[0].keys()
		wb2 = load_workbook(in_filename)
		ws2 = wb2.active
		row = 2
		column = 1
		for student in students:
			for header in headers_list:
				# if student[header] != '' and student[header] is not None:
				c = ws2.cell(row=row, column=column)
				c.value = student[header]
				column += 1
			column = 1
			row += 1
		wb2.save(out_filename)

	def _writejson(self, students: list, filename: str):
		"""write dict list to a json file"""
		#jsonobj = JsonObject(students=students)
		with open(filename, 'w') as fp:
			json.dump(students, fp)
